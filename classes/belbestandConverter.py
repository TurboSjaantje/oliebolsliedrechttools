import openpyxl
import re

class Order:
    def __init__(self, order_number, address, city, postal_code, phone_number, oliebollen, appelbeignets, gift, total_price, paid,
                 delivery_address, delivery_postal_code, delivery_city, comments, date, name, bezorgen):
        self.order_number = order_number
        self.address = address
        self.city = city
        self.postal_code_numbers = "".join(re.findall(r'\d+', postal_code)) if postal_code else ""
        self.postal_code_letters = postal_code[-2:] if postal_code else ""
        self.phone_number = phone_number
        self.oliebollen = oliebollen
        self.appelbeignets = appelbeignets
        self.gift = gift
        self.total_price = total_price
        self.paid = paid
        self.delivery_address = delivery_address
        self.delivery_postal_code = delivery_postal_code
        self.delivery_city = delivery_city
        self.comments = comments
        self.opmerkingen = date
        self.name = name
        self.bezorgen = bezorgen

    def __repr__(self):
        return f"Order({self.order_number}, {self.address}, {self.city}, {self.postal_code_numbers}, {self.phone_number}, {self.oliebollen}, {self.appelbeignets}, {self.gift}, {self.total_price}, {self.paid}, {self.delivery_address}, {self.delivery_postal_code}, {self.delivery_city}, {self.comments}, {self.date})"

    def __str__(self):
        return f"Order({self.order_number}, {self.address}, {self.city}, {self.postal_code_numbers}, {self.phone_number}, {self.oliebollen}, {self.appelbeignets}, {self.gift}, {self.total_price}, {self.paid}, {self.delivery_address}, {self.delivery_postal_code}, {self.delivery_city}, {self.comments}, {self.date})"


class belbestandConverter:
    @staticmethod
    def process_excel_file(file_path, output_file_path):
        def find_column(headers, keyword):
            for name, idx in headers.items():
                if re.search(re.escape(keyword), name, re.IGNORECASE):  # Case-insensitive match
                    return idx
            print(f"Warning: Column '{keyword}' not found!")
            return None

        print('Starting to process the file...')

        # Load the workbook
        workbook = openpyxl.load_workbook(file_path)
        orders = []

        # Loop through all sheets in the workbook
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Extract headers
            headers = {cell.value: idx for idx, cell in enumerate(sheet[1]) if cell.value}
            print("Headers:", headers)

            # Locate necessary columns in the Excel file
            order_number_col = find_column(headers, "Ordernummer")
            address_col = find_column(headers, "Afleveradres")
            city_col = find_column(headers, "Afleverplaats")
            postal_code_col = find_column(headers, "Afleverpostcode")
            phone_number_col = find_column(headers, "Telefoon")
            oliebollen_col_5 = find_column(headers, "Oliebollen (5 stuks)") + 1
            oliebollen_col_10 = find_column(headers, "Oliebollen (10 stuks)") + 1
            oliebollen_col_25 = find_column(headers, "Oliebollen (25 stuks)") + 1
            oliebollen_col_50 = find_column(headers, "Oliebollen (50 stuks)") + 1
            appelbeignets_col_5 = find_column(headers, "Appelbeignets (5 stuks)") + 1
            appelbeignets_col_10 = find_column(headers, "Appelbeignets (10 stuks)") + 1
            gift_col = find_column(headers, "Gift")
            total_price_col = find_column(headers, "Totaal")
            paid_col = find_column(headers, "Betaald")
            comments_col = find_column(headers, "Opmerkingen")
            opmerkingen_col = find_column(headers, "Opmerkingen")
            name_col = find_column(headers, "Naam")
            bezorgen_col = find_column(headers, "Bezorgen?")

            # Check for missing columns
            if any(col is None for col in [order_number_col, address_col, city_col, postal_code_col, phone_number_col]):
                print("Error: Some essential columns are missing!")
                return

            all_rows = list(sheet.iter_rows(values_only=True))

            for row in all_rows[1:]:
                def safe_value(value):
                    try:
                        return float(value) if value not in (None, "") else 0
                    except ValueError:
                        return 0

                oliebollen = (
                        safe_value(row[oliebollen_col_5]) * 5 +
                        safe_value(row[oliebollen_col_10]) * 10 +
                        safe_value(row[oliebollen_col_25]) * 25 +
                        safe_value(row[oliebollen_col_50]) * 50
                )
                appelbeignets = (
                        safe_value(row[appelbeignets_col_5]) * 5 +
                        safe_value(row[appelbeignets_col_10]) * 10
                )

                gift = safe_value(row[gift_col])
                total_price = safe_value(row[total_price_col])
                paid = row[paid_col] if row[paid_col] in ["Ja", "ja", "yes", "Yes"] else "Nee"

                delivery_address = row[address_col]
                delivery_postal_code = row[postal_code_col]
                delivery_city = row[city_col]
                comments = row[comments_col] if comments_col is not None else ""
                order_number = row[order_number_col]
                opmerkingen_col_value = row[opmerkingen_col] if opmerkingen_col is not None else ""

                bezorgen_var = "Ophalen" if row[bezorgen_col] == 0 else "Lokale bezorging"

                order = Order(
                    order_number=order_number,
                    address=row[address_col],
                    city=row[city_col],
                    postal_code=row[postal_code_col],
                    phone_number=row[phone_number_col],
                    oliebollen=oliebollen,
                    appelbeignets=appelbeignets,
                    gift=gift,
                    total_price=total_price,
                    paid=paid,
                    delivery_address=delivery_address,
                    delivery_postal_code=delivery_postal_code,
                    delivery_city=delivery_city,
                    comments=comments,
                    date=opmerkingen_col_value,
                    name=row[name_col],
                    bezorgen=bezorgen_var
                )
                orders.append(order)

        output_workbook = openpyxl.Workbook()
        output_sheet = output_workbook.active
        output_sheet.title = "Processed Orders"
        headers = ["Ordernummer", "Totaalprijs", "Betaald", "Oliebollen", "Appelbeignets", "Naam", "Ophalen/Bezorgen", "Opmerkingen", "Datum", "Tijd",
                   "Factuuradres", "Leveradres", "Factuurpostcode", "Leverpostcode", "Factuurplaats", "Leverplaats", "Telefoonnummer", "Bezorgkosten", "Gift"]
        output_sheet.append(headers)

        for order in orders:
            output_sheet.append([
                order.order_number,
                order.total_price,
                order.paid,
                order.oliebollen,
                order.appelbeignets,
                order.name,
                order.bezorgen,
                order.opmerkingen,
                "",  # Assuming date is formatted properly
                "9:00 - 17:00",  # Time (if available separately)
                order.address,
                order.delivery_address,
                f"{order.postal_code_numbers} {order.postal_code_letters}",
                order.delivery_postal_code,
                order.city,
                order.delivery_city,
                order.phone_number,
                "0.00",  # Bezorgkosten (if applicable)
                order.gift
            ])

        output_workbook.save(output_file_path)
