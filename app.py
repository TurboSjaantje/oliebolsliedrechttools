import math
from collections import defaultdict
from math import isnan

import requests
from flask import Flask, render_template, request, redirect, flash, url_for, send_file, jsonify, abort
import json
import os
import subprocess
import threading
import pandas as pd
import time
import re
import openpyxl
from threading import Lock

from geopy.distance import geodesic
from openpyxl.workbook import Workbook
from classes.belbestandConverter import belbestandConverter
from classes.vrp import VRP, Point


class Order:
    def __init__(self, ordernummer, totaalprijs, betaald, oliebollen, appelbeignets, naam, ophalen_of_bezorgen, datum,
                 tijd, factuuradres, leveradres, factuurpostcode, leverpostcode, factuurplaats, leverplaats,
                 telefoonnummer, bezorgkosten, gift, lat, lon):
        self.ordernummer = ordernummer
        self.totaalprijs = totaalprijs
        self.betaald = betaald
        self.oliebollen = oliebollen
        self.appelbeignets = appelbeignets
        self.naam = naam
        self.ophalen_of_bezorgen = ophalen_of_bezorgen
        self.datum = datum
        self.tijd = tijd
        self.factuuradres = factuuradres
        self.leveradres = leveradres
        self.factuurpostcode = factuurpostcode
        self.leverpostcode = leverpostcode
        self.factuurplaats = factuurplaats
        self.leverplaats = leverplaats
        self.telefoonnummer = telefoonnummer
        self.bezorgkosten = bezorgkosten
        self.gift = gift
        self.lat = lat
        self.lon = lon

    def __str__(self):
        return f"Order #{self.ordernummer} - {self.naam}\nTotal Price: €{self.totaalprijs}\nPaid: {self.betaald}\nOliebollen: {self.oliebollen}, Appelbeignets: {self.appelbeignets}\nPickup/Delivery: {self.ophalen_of_bezorgen}\nDate: {self.datum} {self.tijd}\nInvoice Address: {self.factuuradres} {self.factuurpostcode} {self.factuurplaats}\nDelivery Address: {self.leveradres} {self.leverpostcode} {self.leverplaats}\nPhone: {self.telefoonnummer}\nDelivery Cost: €{self.bezorgkosten}\nGift: {self.gift}"


app = Flask(__name__)
app.secret_key = 'supersecretkey'

# File paths for storing settings
BELBESTAND_FILE = 'belbestand_settings.json'
SHOPIFY_FILE = 'shopify_settings.json'

# Directory for uploaded files
UPLOAD_FOLDER = './uploads'
OUTPUT_FOLDER = './output'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Variables to track uploaded files
uploaded_files = {'file1': None, 'file2': None, 'file3': None}

# Global variables to track merge status and errors
merge_status = {'running': False, 'output_file': None}
error_status = {'error': None}
error_lock = Lock()  # Thread-safe lock for error_status


def run_merge_script(file1, file2, output_file):
    """Simulate running an external Python script."""
    try:
        merge_status['running'] = True  # Mark the task as running
        # Simulate a long-running task
        print("Starting merge process...")
        time.sleep(5)  # Replace this with the actual script logic
        # Write a result file (simulate merged content)
        with open(output_file, 'w') as f:
            f.write(f"Merged content of:\n{file1}\n{file2}\n")
        merge_status['output_file'] = output_file  # Update the output file path
        print("Merge process completed!")
    except Exception as e:
        print(f"Error during merge: {e}")
    finally:
        merge_status['running'] = False  # Ensure task is marked as not running


# Function to save settings to a JSON file
def save_settings_to_file(filename, data):
    with open(filename, 'w') as file:
        json.dump(data, file, indent=4)


# Function to load settings from a JSON file
def load_settings_from_file(filename):
    try:
        with open(filename, 'r') as file:
            return json.load(file)
    except FileNotFoundError:
        return {}


@app.route('/')
def index():
    files_uploaded = all(uploaded_files.values())  # True if both files are uploaded
    return render_template('tools.html', files_uploaded=files_uploaded, uploaded_files=uploaded_files)


@app.route('/upload', methods=['POST'])
def upload_files():
    global uploaded_files
    # Handle file1 upload (Bestand Belavond)
    if 'file1' in request.files and request.files['file1'].filename:
        file1 = request.files['file1']
        file1_path = os.path.join(UPLOAD_FOLDER, file1.filename)
        file1.save(file1_path)
        uploaded_files['file1'] = file1.filename

    # Handle file2 upload (Bestand Shopify)
    if 'file2' in request.files and request.files['file2'].filename:
        file2 = request.files['file2']
        file2_path = os.path.join(UPLOAD_FOLDER, file2.filename)
        file2.save(file2_path)
        uploaded_files['file2'] = file2.filename

    flash("Files uploaded successfully.", "success")
    return redirect(url_for('index'))


@app.route('/upload-routemaker', methods=['POST'])
def upload_routemaker():
    global uploaded_files
    if 'file3' in request.files and request.files['file3'].filename:
        file3 = request.files['file3']
        try:
            file3_path = os.path.join(UPLOAD_FOLDER, file3.filename)
            file3.save(file3_path)
            uploaded_files['file3'] = file3.filename
            flash(f"File '{file3.filename}' uploaded successfully.", "success")
        except Exception as e:
            flash(f"Failed to upload file: {str(e)}", "error")
    else:
        flash("No file selected for upload.", "warning")
    return redirect(url_for('routemaker'))


@app.route('/routemaker', methods=['GET'])
def routemaker():
    file_uploaded = uploaded_files['file3']
    print(file_uploaded, uploaded_files['file3'])
    return render_template('routemaker.html', file_uploaded=file_uploaded, uploaded_file=uploaded_files['file3'])


@app.route('/remove/<file_key>', methods=['POST'])
def remove_file(file_key):
    global uploaded_files
    if file_key in uploaded_files and uploaded_files[file_key]:
        file_path = os.path.join(UPLOAD_FOLDER, uploaded_files[file_key])
        if os.path.exists(file_path):
            os.remove(file_path)
        uploaded_files[file_key] = None
        flash(f"{file_key.capitalize()} removed successfully.", "success")
    else:
        flash("No file to remove.", "error")
    return redirect(url_for('index'))


@app.route('/settings', methods=['GET'])
def settings():
    # Load settings for both belbestand and shopify
    belbestand_settings = load_settings_from_file(BELBESTAND_FILE)
    shopify_settings = load_settings_from_file(SHOPIFY_FILE)
    return render_template('settings.html', belbestand=belbestand_settings, shopify=shopify_settings)


@app.route('/settings/belbestand', methods=['POST'])
def belbestand():
    # Process belbestand form data
    belbestand_settings = request.form.to_dict()
    save_settings_to_file(BELBESTAND_FILE, belbestand_settings)
    flash("Belbestand settings saved successfully!", "success")
    return redirect(url_for('settings'))


@app.route('/settings/shopifybestand', methods=['POST'])
def shopifybestand():
    # Process shopify form data
    shopify_settings = request.form.to_dict()
    save_settings_to_file(SHOPIFY_FILE, shopify_settings)
    flash("Shopify settings saved successfully!", "success")
    return redirect(url_for('settings'))


@app.route('/merge-files', methods=['POST'])
def merge_files():
    if not all(uploaded_files.values()):
        flash("Both files must be uploaded to merge.", "error")
        return redirect(url_for('index'))

    # Start the merge task
    merge_status['running'] = True
    merge_status['output_file'] = None
    file1 = os.path.join(UPLOAD_FOLDER, uploaded_files['file1'])
    file2 = os.path.join(UPLOAD_FOLDER, uploaded_files['file2'])
    output_file = os.path.join(OUTPUT_FOLDER, 'merged_result.txt')

    # Run the merge script in a separate thread
    thread = threading.Thread(target=run_merge_script, args=(file1, file2, output_file))
    thread.start()

    return jsonify({'status': 'started'})


@app.route('/download-merged', methods=['GET', 'HEAD'])
def download_merged():
    if not merge_status['output_file']:
        return ('', 404)  # Return a 404 if the file is not ready
    if request.method == 'HEAD':
        return ('', 200)  # Indicate the file is ready
    return send_file(merge_status['output_file'], as_attachment=True)


@app.route('/normalize-shopify/<year_key>', methods=['POST'])
def normalize_shopify_tool(year_key):
    if not uploaded_files['file2']:
        flash("No Shopify file uploaded to normalize.", "error")
        return jsonify({'status': 'error', 'message': 'No Shopify file uploaded.'})

    # Paths for input and output
    input_file = os.path.join(UPLOAD_FOLDER, uploaded_files['file2'])
    output_file = os.path.join(OUTPUT_FOLDER, 'normalized_shopify.xlsx')

    # Start normalization task in a separate thread
    merge_status['running'] = True
    merge_status['output_file'] = None
    thread = threading.Thread(target=normalize_shopify, args=(input_file, output_file, year_key))
    thread.start()

    return jsonify({'status': 'started'})


@app.route('/normalize-belbestand', methods=['POST'])
def normalize_belbestand_tool():
    if not uploaded_files['file1']:
        flash("No belbestand uploaded to normalize.", "error")
        return jsonify({'status': 'error', 'message': 'No belbestand uploaded.'})

    # Paths for input and output
    input_file = os.path.join(UPLOAD_FOLDER, uploaded_files['file1'])
    output_file = os.path.join(OUTPUT_FOLDER, 'normalized_belbestand.xlsx')

    # Start normalization task in a separate thread
    merge_status['running'] = True
    merge_status['output_file'] = None
    thread = threading.Thread(target=normalize_belbestand, args=(input_file, output_file))
    thread.start()

    return jsonify({'status': 'started'})


@app.route('/check-normalized-shopify', methods=['GET', 'HEAD'])
def check_normalized_shopify():
    if not merge_status['output_file']:
        return ('', 404)  # File not ready
    if request.method == 'HEAD':
        return ('', 200)  # File ready for download
    return send_file(merge_status['output_file'], as_attachment=True)


@app.route('/check-normalized-belbestand', methods=['GET', 'HEAD'])
def check_normalized_belbestand():
    if not merge_status['output_file']:
        return ('', 404)  # File not ready
    if request.method == 'HEAD':
        return ('', 200)  # File ready for download
    return send_file(merge_status['output_file'], as_attachment=True)


def normalize_shopify(input_file, output_file, year):
    try:
        json_file = "./shopify_settings.json"
        with open(json_file, "r") as file:
            column_mapping = json.load(file)

        reversed_mapping = {v: k for k, v in column_mapping.items()}
        df = pd.read_csv(input_file)

        # Filter rows and rename columns
        order_rows = df[df['Name'].str.contains(r'^#\d+', na=False, flags=re.IGNORECASE)]
        columns_to_keep = list(column_mapping.values())
        filtered_rows = order_rows[order_rows['Created at'].str.contains(year, na=False, flags=re.IGNORECASE)]

        if filtered_rows.empty:
            raise ValueError("No orders found for the specified year.")

        filtered_rows = filtered_rows[columns_to_keep].rename(columns=reversed_mapping)

        # Initialize new columns
        filtered_rows.loc[:, 'Oliebollen'] = 0
        filtered_rows.loc[:, 'Appelbeignets'] = 0

        # Create Excel workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders"

        # Append header row with rearranged columns
        header = ["Ordernummer", "Totaalprijs", "Betaald", "Oliebollen", "Appelbeignets",
                  "Naam", "Ophalen/Bezorgen", "Opmerkingen", "Datum", "Tijd", "Factuuradres", "Leveradres",
                  "Factuurpostcode", "Leverpostcode", "Factuurplaats", "Leverplaats",
                  "Telefoonnummer", "Bezorgkosten", "Gift"]
        ws.append(header)

        # Dictionary to track row locations for each Ordernummer
        row_map = {}

        for index, row in filtered_rows.iterrows():
            current_ordernummer = row['Ordernummer']

            # Initialize variables for the current row
            aantal_oliebollen = 0
            aantal_appelbeignets = 0
            gift = 0
            betaald = "Ja" if row['Betaald'] == "paid" else "Nee"

            # Parse the Bezorging field
            ophalen_bezorgen, datum, tijd = "N/A", "N/A", "N/A"
            if pd.notnull(row['Bezorging']):
                try:
                    bezorging_parts = row['Bezorging'].split("|")
                    ophalen_bezorgen = bezorging_parts[0].split(":")[1].strip() if len(bezorging_parts) > 0 else "N/A"
                    datum = bezorging_parts[1].split(":")[1].strip() if len(bezorging_parts) > 1 else "N/A"
                    tijd = bezorging_parts[2].split("Time:")[1].strip() if len(bezorging_parts) > 2 else "N/A"
                except Exception as e:
                    print(f"Error parsing Bezorging field: {e}")

            # Your original logic to determine amounts
            if "Combi" in row['Productkolom']:
                split_row = row['Productkolom'].split(" ")
                aantal_oliebollen = int(split_row[1]) * int(row['Kwantiteitkolom'])
                aantal_appelbeignets = int(split_row[4]) * int(row['Kwantiteitkolom'])
            elif "Oliebollen" in row['Productkolom']:
                match = re.findall(r'\d+', row['Productkolom'])
                if match:
                    aantal_oliebollen = int(match[0]) * int(row['Kwantiteitkolom'])
            elif "Appelbeignets" in row['Productkolom']:
                match = re.findall(r'\d+', row['Productkolom'])
                if match:
                    aantal_appelbeignets = int(match[0]) * int(row['Kwantiteitkolom'])
            elif "Tip" in row['Productkolom']:
                gift = int(row['Gift']) if pd.notnull(row['Gift']) else 0

            if current_ordernummer not in row_map:
                # New order: create a new row in the workbook
                current_row = [
                    row['Ordernummer'], row['Totaalprijs'], betaald, aantal_oliebollen, aantal_appelbeignets,
                    row['Naam'], ophalen_bezorgen, "", datum, tijd, row['Factuuradres'], row['Leveradres'],
                    row['Factuurpostcode'], row['Leverpostcode'], row['Factuurplaats'], row['Leverplaats'],
                    row['Telefoonnummer'], row['Bezorgkosten'], gift
                ]
                ws.append(current_row)
                row_map[current_ordernummer] = ws.max_row  # Track the row index for this order
            else:
                # Existing order: update the existing row
                target_row = row_map[current_ordernummer]

                # Retrieve current values
                current_oliebollen = int(ws.cell(row=target_row, column=4).value or 0)  # Oliebollen
                current_appelbeignets = int(ws.cell(row=target_row, column=5).value or 0)  # Appelbeignets
                current_gift = int(ws.cell(row=target_row, column=18).value or 0)  # Gift

                # Update values for the current order
                ws.cell(row=target_row, column=4).value = current_oliebollen + aantal_oliebollen
                ws.cell(row=target_row, column=5).value = current_appelbeignets + aantal_appelbeignets
                ws.cell(row=target_row, column=18).value = current_gift + gift

        # Save the workbook
        wb.save(output_file)
        merge_status['output_file'] = output_file
        print("Shopify file normalized successfully!")
    except Exception as e:
        # Safely update the error_status dictionary
        with error_lock:
            error_status['error'] = str(e)
            print(f"Error normalizing Shopify file: {e}")  # Log the error instead

        return redirect(url_for('tools'))
    finally:
        merge_status['running'] = False


def normalize_belbestand(input_file, output_file):
    try:
        print('test')
        belbestandConverter.process_excel_file(input_file, output_file)
        print('tester')

        # Save the workbook
        merge_status['output_file'] = output_file
        print("Belbestand file normalized successfully!")
    except Exception as e:
        # Safely update the error_status dictionary
        with error_lock:
            error_status['error'] = str(e)
            print(f"Error normalizing belbestand: {e}")  # Log the error instead
    finally:
        merge_status['running'] = False

        # Instead of redirecting here, we'll set the status or send a response directly
        # to inform the client that the process has completed or failed.


def split_postcode(postcode):
    # Ensure postcode is a string and handle empty or None
    if not postcode:
        return None, None
    postcode = str(postcode).replace(" ", "")  # Convert to string and remove spaces
    match = re.match(r"(\d{4})([a-zA-Z]{2})", postcode)
    return match.groups() if match else (None, None)


def get_lat_lon(order: Order):
    api_key = "N4MmxXhw-NAahWKSYbe6WaQpGeOWXH1kz_B0ayWdTNI"
    try:
        address = ""

        # Check if order fields are NaN or empty, handling both float NaN and None
        def is_nan(value):
            return value is None or (isinstance(value, float) and math.isnan(value))

        # Check if the 'leveradres', 'leverplaats' are not NaN or empty
        if not is_nan(order.leveradres) and not is_nan(order.leverplaats):
            postcode_parts = split_postcode(order.leverpostcode)
            if postcode_parts != (None, None):
                address = f"{order.leveradres}, {order.leverplaats}, {postcode_parts[0]} {postcode_parts[1]}"
            else:
                address = f"{order.leveradres}, {order.leverplaats}, Invalid postcode"
        else:
            postcode_parts = split_postcode(order.factuurpostcode)
            if postcode_parts != (None, None):
                address = f"{order.factuuradres}, {order.factuurplaats}, {postcode_parts[0]} {postcode_parts[1]}"
            else:
                address = f"{order.factuuradres}, {order.factuurplaats}, Invalid postcode"

        # print("Address: ", address)

        # Make sure address is a string before using it in the API call
        address = str(address).strip()

        # Geocoding URL with address query
        url = f"https://geocode.search.hereapi.com/v1/geocode?q={address}&apiKey={api_key}"

        # print(f"Geocoding address: {address}")

        # # Make the HTTP GET request
        response = requests.get(url)
        response.raise_for_status()  # Raise an HTTPError if the response is not 200

        # Parse the JSON response
        data = response.json()

        if "items" in data and len(data["items"]) > 0:
            # Check if "access" is present in the first result
            if "access" in data["items"][0] and len(data["items"][0]["access"]) > 0:
                # Get latitude and longitude from the first access point
                access_point = data["items"][0]["access"][0]
                order.lat = access_point["lat"]
                order.lon = access_point["lng"]
            else:
                print(f"No 'access' points found for address: {address}")
                order.lat, order.lon = None, None  # Set None if no access points
            return order
        else:
            print(f"No geocoding results found for address: {address}")
            return order  # Return the order object even if no lat/lon found
    except Exception as e:
        print(f"Error geocoding address '{address}': {e}")
        return order  # Return the order object to maintain flow even on failure


def generate_routes_vrp(points, num_vehicles):
    vehicle_capacity = len(points) // num_vehicles

    vrp = VRP(points, num_vehicles, vehicle_capacity)

    vrp.greedy_vrp()

    vrp.optimize_routes()

    return vrp.routes


def calculate_distance(coord1, coord2):
    return geodesic(coord1, coord2).km


def distribute_orders(orders, aantal_bezorgers):
    # Sort orders by proximity
    order_coords = [(order, (order.lat, order.lon)) for order in orders]

    # Initialize an empty list to store the routes
    routes = defaultdict(list)

    # Set the maximum orders per driver
    max_orders_per_bezorger = math.ceil(len(orders) / aantal_bezorgers)

    # Keep track of the last order placed with each driver
    last_location = [None] * aantal_bezorgers  # To store the last location for each driver

    # Distribute the orders greedily to the closest available driver
    for order, coords in order_coords:
        # Find the closest driver
        min_distance = float('inf')
        best_driver = None

        for i in range(aantal_bezorgers):
            # If it's the first order for the driver or not exceeding the max limit
            if last_location[i] is None or len(routes[i]) < max_orders_per_bezorger:
                # Calculate distance to last assigned order for the driver
                if last_location[i] is None:
                    distance = 0  # No previous orders for this driver, assign first order
                else:
                    distance = calculate_distance(last_location[i], coords)

                if distance < min_distance:
                    min_distance = distance
                    best_driver = i

        # Assign this order to the best driver
        routes[best_driver].append(order)
        last_location[best_driver] = coords  # Update the driver's last location

    return routes


def generate_routes(input_file, output_file, dagdeel, aantal_bezorgers):
    try:
        print(f"Generating routes with Dagdeel: {dagdeel}, Aantal Bezorgers: {aantal_bezorgers}")

        # Initialize the VRP class
        df = pd.read_excel(input_file, sheet_name='dagdeel ' + str(dagdeel))
        orders = []

        for _, row in df.iterrows():
            order = Order(
                ordernummer=row['Ordernummer'],
                totaalprijs=row['Totaalprijs'],
                betaald=row['Betaald'],
                oliebollen=row['Oliebollen'],
                appelbeignets=row['Appelbeignets'],
                naam=row['Naam'],
                ophalen_of_bezorgen=row['Ophalen/Bezorgen'],
                datum=row['Datum'],
                tijd=row['Tijd'],
                factuuradres=row['Factuuradres'],
                leveradres=row['Leveradres'],
                factuurpostcode=row['Factuurpostcode'],
                leverpostcode=row['Leverpostcode'],
                factuurplaats=row['Factuurplaats'],
                leverplaats=row['Leverplaats'],
                telefoonnummer=row['Telefoonnummer'],
                bezorgkosten=row['Bezorgkosten'],
                gift=row['Gift'],
                lat=None,  # Ensure lat and lon are present in your input
                lon=None
            )
            orders.append(order)

        # Assign lat/lon to orders (if not already done)
        orders = [get_lat_lon(order) for order in orders]

        # Distribute orders among drivers
        routes = distribute_orders(orders, aantal_bezorgers)

        # Write to an Excel file
        with pd.ExcelWriter(output_file) as writer:
            for bezorger, orders_for_bezorger in routes.items():
                # Convert the orders to a DataFrame for the current driver
                orders_data = pd.DataFrame([order.__dict__ for order in orders_for_bezorger])
                sheet_name = f'Bezorger {bezorger + 1}'
                orders_data.to_excel(writer, sheet_name=sheet_name, index=False)

        # Set the route file path in merge_status once done
        merge_status['route_file'] = output_file
        print("Routes generated successfully!")
    except Exception as e:
        print(f"Error generating routes: {e}")
        raise


@app.route('/generate-routes', methods=['POST'])
def generate_routes_endpoint():
    print('testerererere')

    # Ensure the file has been uploaded
    uploaded_file = os.path.join(UPLOAD_FOLDER, uploaded_files.get('file3', ''))

    if not os.path.exists(uploaded_file):
        return jsonify({'status': 'error', 'message': 'No uploaded file found.'}), 400

    # Get the Dagdeel and Aantal Bezorgers inputs
    data = request.get_json()
    dagdeel = int(data['dagdeel'])
    aantal_bezorgers = int(data['aantal_bezorgers'])

    if not dagdeel or not (1 <= dagdeel <= 3):
        return jsonify({'status': 'error', 'message': 'Invalid Dagdeel value.'}), 400

    if not aantal_bezorgers or aantal_bezorgers < 1:
        return jsonify({'status': 'error', 'message': 'Invalid Aantal Bezorgers value.'}), 400

    # Path for output file
    routes_file = os.path.join(OUTPUT_FOLDER, 'generated_routes.xlsx')

    # Start the route generation process
    try:
        print('Starting route generation...')
        thread = threading.Thread(target=generate_routes, args=(uploaded_file, routes_file, dagdeel, aantal_bezorgers))
        thread.start()
        return jsonify({'status': 'started'}), 200
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/download-routes', methods=['HEAD', 'GET'])
def download_routes():
    # Check if the route file exists in the merge_status
    route_file = merge_status.get('route_file')  # Retrieve the file path from merge_status

    if not route_file or not os.path.exists(route_file):
        abort(404)  # File not found

    if request.method == 'HEAD':
        return '', 200  # Indicate that the file exists

    return send_file(route_file, as_attachment=True)


@app.route('/check-errors', methods=['GET'])
def check_errors():
    """API endpoint to check for errors."""
    with error_lock:
        error = error_status.get('error')
        if error:
            error_status['error'] = None  # Clear the error after reporting it
            return jsonify({'status': 'error', 'message': error})
    return jsonify({'status': 'ok'})


if __name__ == '__main__':
    app.run(debug=True)
